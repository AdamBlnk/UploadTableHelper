from PyQt5.QtWidgets import QVBoxLayout, QPushButton, QScrollArea, QDialog, QTextEdit, QLabel, QLineEdit, QFormLayout, QMessageBox
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class scriptData:
    mySheet, cellEdit, phraseEdit, textEdit, scrollArea, scriptPath = "", "", "", "", "", "", 
    
    def __init__(self, mySheet, cellEdit, cellEditVar, phraseEdit, textEdit, scrollArea, scriptPath, workbook):
        self.mySheet = mySheet          # The worksheet
        self.cellEdit = cellEdit        # Column entered by the user
        self.cellEditVar = cellEditVar  # This holds the data typed into the cellEdit
        self.phraseEdit = phraseEdit    # The user entered phrase
        self.textEdit = textEdit        # Stores the text that will be displayed in scrollArea
        self.scrollArea = scrollArea    # Displays the changes
        self.scriptPath = scriptPath    # Path to the folder that holds the script
        self.workbook = workbook        # Whole workbook




# Depending on confirmation, this will return a list of each row with the user's text at the beginning
def getText(frontOfCell, confirmation):
    # textList will hold what each row will say with the user input
    textList = []
    
    for i in range(1, frontOfCell.mySheet.max_row + 1, 1):
        index = frontOfCell.mySheet.cell(i, frontOfCell.cellEditVar).value

        # if confirmation is None, this means we are previewing the changes, include row numbers
        if confirmation is None:

            # textList is the :03d, then phrase + index
            if (index is not None):
                textList.append("{:03d}: {}{}".format(i, frontOfCell.phraseEdit.text(), index))
        
        # if confirmation is true, that means we are confirming the final changes, so we drop the row numbers from the string
        else:
            if (index is not None):
                textList.append(f"{frontOfCell.phraseEdit.text()}{index}")

    return textList


def PreviewText(frontOfCell, confirmation):
    # Get text returns a list of the rows, format determined by confirmation
    textList = getText(frontOfCell, confirmation)

    # text becomes the list but each element is joined with a \n
    # then textEdit becomes the text string, and the textEdit is put into the scrollArea
    text = '\n'.join(textList)
    frontOfCell.textEdit.setPlainText(text)
    frontOfCell.scrollArea.setWidget(frontOfCell.textEdit)


def confirmChange (frontOfCell):
    for i in range(1, frontOfCell.mySheet.max_row + 1, 1):
        frontOfCell.mySheet.cell(i, frontOfCell.cellEditVar).value = frontOfCell.phraseEdit.text() + frontOfCell.mySheet.cell(i, frontOfCell.cellEditVar).value
    
    frontOfCell.workbook.save(frontOfCell.scriptPath)


# The 'confirmed' var is either None or True
def confirmationDialog(frontOfCell, confirmation):
    msgBox = QMessageBox()
    msgBox.setIcon(QMessageBox.Question)
    msgBox.setWindowTitle("Confirmation")
    msgBox.setText("Are you sure you want to proceed?\nChanges made cannot be undone")
    msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    msgBox.setDefaultButton(QMessageBox.No)
    result = msgBox.exec_()

    if result == QMessageBox.Yes:
        handleButtons(frontOfCell, confirmation)
    else:
        print("User selected no to confirming\nClosing confirmation window")
        

def handleButtons(frontOfCell, confirmation):
    # cellEdit provides user input, cellEditVar handles the manipulation
    frontOfCell.cellEditVar = frontOfCell.cellEdit.text()

    # This if statement changes whatever was written into cellText into a number
    # No matter what the user enters, it is a string
    # That string is either '2', which needs to be cast into an int, or 'B', which needs to be turned into the int 2
    if type(frontOfCell.cellEditVar) is str:
        if frontOfCell.cellEditVar.isnumeric():
            frontOfCell.cellEditVar = int(frontOfCell.cellEditVar) 
        else:
            frontOfCell.cellEditVar = column_index_from_string(frontOfCell.cellEditVar)   

    if confirmation == None:
        # Call the PreviewText function with the required arguments
        PreviewText(frontOfCell, confirmation)
    else:
        # confirm button pressed
        confirmChange(frontOfCell)


def AddTextToCellBeginning(excelPath):
    # load the workbook and get the first sheet (Assets sheet)
    wb = load_workbook(excelPath)
    mySheet = wb.worksheets[0]

    # New window beginnings
    dialog = QDialog()
    dialog.setWindowTitle("Add Text To The Beginning of the Cell")
    
    # Buttons
    previewButton = QPushButton("Preview Changes")
    previewButton.setFixedSize(100, 25)
    confirmButton = QPushButton("Confirm")
    confirmButton.setFixedSize(100, 25)

    # Labels
    cellLabel = QLabel("Cell column:")
    phraseLabel = QLabel("Phrase to insert:")

    # User input boxes (QLineEdit)
    cellEdit = QLineEdit()
    cellEdit.setText('4')
    phraseEdit = QLineEdit()
    phraseEdit.setText("")

    # Scroll area 
    scrollArea = QScrollArea()
    textEdit = QTextEdit()
    textEdit.setReadOnly(True)
    scrollArea.setWidgetResizable(True)
    scrollArea.setWidget(textEdit)

    

    # Grid to hold the Labels, Edits, and Buttons
    grid = QFormLayout()
    grid.addRow(cellLabel, cellEdit)
    grid.addRow(phraseLabel, phraseEdit)
    grid.addRow(previewButton, confirmButton)
    grid.setAlignment(confirmButton, Qt.AlignRight)

    # Layout boxes
    vbox = QVBoxLayout()
    vbox.setAlignment(Qt.AlignTop)

    vbox.addLayout(grid)
    vbox.addWidget(scrollArea)

    # Set the layout to the vbox
    dialog.setLayout(vbox)

    # using the declarations above, create an object that holds all the information
    frontOfCell = scriptData(mySheet, cellEdit, "", phraseEdit, textEdit, scrollArea, excelPath, wb)

    # Connect to the appropriate methods on button clicks
    # The last variable is for the handleButtons method as a way to reuse the method
    # If its None, that makes it preview the changes, if True, it makes the changes
    previewButton.clicked.connect(lambda: handleButtons (frontOfCell, None))
    confirmButton.clicked.connect(lambda: confirmationDialog(frontOfCell, True))

    # Set the initial window size
    dialog.resize(400, 200)  # width, height
    dialog.exec_()